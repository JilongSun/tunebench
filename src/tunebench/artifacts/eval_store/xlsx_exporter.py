"""评测 CSV 到 XLSX 的汇总导出器。"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import EvalReportExportResult


class XlsxEvalReportExporter:
    """将多个 CSV 评测产物汇总为一个 XLSX。"""

    def _read_csv_rows(self, csv_path: Path) -> tuple[list[str], list[list[str]]]:
        with csv_path.open("r", encoding="utf-8", newline="") as file_obj:
            reader = csv.reader(file_obj)
            rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]

    def _style_worksheet(self, worksheet) -> None:  # type: ignore[no-untyped-def]
        """为导出的评测 sheet 增加基础可读性格式。"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))

            adjusted_width = min(max(max_length + 2, 10), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_from_csv_sources(
        self,
        csv_sources: list[Path],
        output_path: Path,
        optional_sources: set[Path] | None = None,
    ) -> EvalReportExportResult:
        """将指定 CSV 列表汇总导出为一个 XLSX。"""
        workbook = Workbook()
        active_sheet = workbook.active
        if active_sheet is not None:
            workbook.remove(active_sheet)

        warnings: list[str] = []
        created_sheet_count = 0
        optional_sources = optional_sources or set()

        for csv_path in csv_sources:
            if not csv_path.exists():
                if csv_path in optional_sources:
                    warnings.append(f"缺少可选文件: {csv_path.name}")
                    continue
                raise FileNotFoundError(f"缺少必要评测文件: {csv_path}")

            header, data_rows = self._read_csv_rows(csv_path)
            worksheet = workbook.create_sheet(title=csv_path.stem)
            created_sheet_count += 1
            if header:
                worksheet.append(header)
            for row in data_rows:
                worksheet.append(row)
            self._style_worksheet(worksheet)

        if created_sheet_count == 0:
            return EvalReportExportResult(
                output_path=None,
                warnings=["没有可导出的 CSV 文件，跳过 XLSX 汇总导出。"],
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return EvalReportExportResult(output_path=output_path, warnings=warnings)